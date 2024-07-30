import csv
import json
import sys
import pandas as pd
from openpyxl import load_workbook
from io import StringIO

def xlsx_to_simplified_csv(filename):
    # read in excel file from filename
    wb = load_workbook(filename)
    ws = wb.active
    # write to new csv file
    with open('emissions1038-0610-0614-day_auto.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)    
    # clean up csv file to rejoin split headings
    with open('emissions1038-0610-0614-day_auto.csv', 'r') as f:
        rows = f.readlines()
    with open('emissions1038-0610-0614-day_auto_cleaned.csv', 'w') as f:
        for row in rows[:4]:
            f.write(row)  
        detailed_headings = rows[4].replace('\n',' ')
        for row_index in range(5,78):
            detailed_headings += rows[row_index].replace('\n',' ')
        f.write(detailed_headings)
        for row in rows[78:]:
            f.write(row)
    data_dict = {}
    with open('emissions1038-0610-0614-day_auto_cleaned.csv', 'r') as f:
        data_dict['customer'] = f.readline().split(',')[0][-5:].strip()
        line = f.readline().split(',')
        data_dict['data_pulled'] = line[0][13:]
        data_dict['total_secs'] = line[5].strip()
        aws = "Mon,Tue,Wed,Th,Fri, 20:00 to 8:00"
        data_dict['analysis_window'] = f.readline().strip().replace('"','')[15:len(aws)+17].strip()
        machine_lines = []
        line = f.readline()
        top_line_headings = line.strip().split(',')
        line = f.readline()
        sub_headings = line.strip().split(',')
        while "Total(MB)" not in line:
            machine_lines.append(line.strip().split(","))
            line = f.readline()
    # machine_dict = {}
    # for machine_index in range(len(machine_lines)):
    #     for i in range(len(machine_lines[machine_index])):
    #         if top_line_headings[i]:
    #             j = i
    #             machine_dict[top_line_headings[i]] = {}
    #             while not top_line_headings[j]:
    #                 machine_dict[top_line_headings[i]][sub_headings[j]] = machine_lines[machine_index][j]
    #                 j += 1  
    #             i = j
    # for i in range(len(machine_lines[1:])+1):
    #     print(i)     
    #     for k in list(machine_dict.keys()):
    #         if k == 'Host Name':
    #             machine_dict[k] = {'Host Name': machine_lines[i][0]}
    #         if k == 'Model':
    #             del machine_dict[k]
    #         if k == 'CPU %Utilization':
    #             """while total MB not in line, we should iterate through the machine lines in range 2,14 as the value """
    #             machine_dict[k] = {sub_headings[key]: machine_lines[i][key] for key in range(2, 14)}
    #         if k == 'Memory %Utilization':
    #             machine_dict[k] = {sub_headings[key]: machine_lines[i][key] for key in range(6, 21)}
    #         if k == 'All Network Traffic':
    #             machine_dict[k] = {sub_headings[key]: machine_lines[i][key] for key in range(21, 29)}
    #         if k == 'PCoIP Statistics':
    #             machine_dict[k] = {sub_headings[key]: machine_lines[i][key] for key in range(29, 43)}
    #         if k == 'NVIDIA %Utilization':
    #             machine_dict[k] = {sub_headings[key]: machine_lines[i][key] for key in range(43, 51)}
    #         if k == 'Local Disk (workstation internal drives)':
    #             machine_dict[k] = {sub_headings[key]: machine_lines[i][key] for key in range(51, 66)}

    #     Initialize machine_dict with lists to hold data for each row
    machine_dict = {
        'Host Name': {'Host Name': []},
        'CPU %Utilization': {sub_headings[key]: [] for key in range(2, 14)},
        'Memory %Utilization': {sub_headings[key]: [] for key in range(16, 21)},
        'All Network Traffic': {sub_headings[key]: [] for key in range(21, 29)},
        'PCoIP Statistics': {sub_headings[key]: [] for key in range(29, 43)},
        'NVIDIA %Utilization': {sub_headings[key]: [] for key in range(43, 51)},
        'Local Disk (workstation internal drives)': {sub_headings[key]: [] for key in range(51, 66)}
    }

    # Iterate through each row and populate the machine_dict
    for i in range(1, len(machine_lines)):
        for k in machine_dict.keys():
            if k == 'Host Name':
                machine_dict[k]['Host Name'].append(machine_lines[i][0])
            elif k == 'CPU %Utilization':
                for key in range(2, 14):
                    machine_dict[k][sub_headings[key]].append(machine_lines[i][key])
            elif k == 'Memory %Utilization':
                for key in range(16, 21):
                    machine_dict[k][sub_headings[key]].append(machine_lines[i][key])
            elif k == 'All Network Traffic':
                for key in range(21, 29):
                    machine_dict[k][sub_headings[key]].append(machine_lines[i][key])
            elif k == 'PCoIP Statistics':
                for key in range(29, 43):
                    machine_dict[k][sub_headings[key]].append(machine_lines[i][key])
            elif k == 'NVIDIA %Utilization':
                for key in range(43, 51):
                    machine_dict[k][sub_headings[key]].append(machine_lines[i][key])
            elif k == 'Local Disk (workstation internal drives)':
                for key in range(51, 66):
                    machine_dict[k][sub_headings[key]].append(machine_lines[i][key])

    return machine_dict, data_dict

    
    
def csv_to_json(csv_filename, as_json=True):
    data_dict = {}
    with open(csv_filename, 'r') as f:
        reader = csv.DictReader(f)
        for machine_row in reader:
            machine_key = machine_row[next(iter(machine_row))]
            data_dict[machine_key] = {}
            for header, value in machine_row.items():
                if header == next(iter(machine_row)):
                    continue
                # key before parentheses
                lower_level_key = header.split('(')[0].strip().replace('\n',' ')
                # key within parentheses
                top_level_key = header.split('(')[-1].split(')')[0].strip().replace('\n',' ')
                if top_level_key not in data_dict[machine_key]:
                    data_dict[machine_key][top_level_key] = {}
                try:
                    value = str(round(float(value),2)) 
                except ValueError:
                    value = value
                if top_level_key != lower_level_key:
                    data_dict[machine_key][top_level_key][lower_level_key] = value
                else:
                    data_dict[machine_key][top_level_key] = value
    if as_json:
        return json.dumps(data_dict)
    return data_dict
    

def flatten(data_dict):
    flat_dict = {}
    for machine_key, machine_data in data_dict.items():
        machine_key = machine_key[6:]
        for top_level_key, top_level_data in machine_data.items():
            if isinstance(top_level_data, dict):
                for lower_level_key, lower_level_data in top_level_data.items():
                    flat_dict[machine_key + '_' + top_level_key + '_' + lower_level_key] = lower_level_data
            else:
                flat_dict[machine_key + '_' + top_level_key] = top_level_data
    return flat_dict

def stringify(flat_dict):
    dict_string = ""
    for k,v in flat_dict.items():
        dict_string += k.replace('_',' ') + " = " + v + "\n"
    # Changed to dict_string return not flat_dict
    dict_string = deabbreviate(dict_string)
    print('dict_string:', dict_string)
    return dict_string

def deabbreviate(sentence: str) -> str:
    abbr_list = {' CPU ': ' central processing unit ', 'GPU': 'graphics processing unit', '%':'Percent ', 'Mem ':'Memory ', 'min ':'minimum', 'max ':'maximum ', 'avg ':'average ', '#': ' number of ', 'mb ':'megabytes ',
                 'oc ': 'occurrences '} 
    sentence = sentence.lower()
    for abbr, full in abbr_list.items():
        sentence = sentence.replace(abbr.lower(), full.lower())
    return sentence



if __name__ == '__main__':
    
    # Doesn't Work
    # Read the Excel file
    excel_file = r'data\1038-0610-0614-evening.xlsx'
    df = pd.read_excel(excel_file, sheet_name='WS-Data')  # Change 'Sheet1' to the name of the sheet you want to read
    print(df)

    # # Save the DataFrame as a CSV file
    # csv_file = r'C:\Users\martha.calder-jones\OneDrive - University College London\UCL_comp_sci\Sustainable_Systems_3\HP_Sus_Sys_3\emissions1038-0610-0614-day_auto_cleaned.csv'
    # df = pd.read_csv(csv_file)
    # print(df)
    
    filename = r"C:\Users\martha.calder-jones\OneDrive - University College London\UCL_comp_sci\Sustainable_Systems_3\HP_Sus_Sys_3\data\1038-0610-0614-evening.xlsx"
    machine_dict, data_dict = xlsx_to_simplified_csv(filename)
    print(data_dict)
    print(machine_dict)
    sys.exit()
    filename = r"C:\Users\martha.calder-jones\OneDrive - University College London\UCL_comp_sci\Sustainable_Systems_3\HP_Sus_Sys_3\data\emissions1038-0610-0614-day.csv"
    
    data_dict_json = csv_to_json(csv_filename, as_json=False)
    print(flatten(data_dict_json))
    #  Added below
    flat_dict = flatten(data_dict_json)
    dict_string = stringify(flat_dict)
   
    # """save to file"""
    # with open('data.json', 'w') as f:
    #     f.write(dict_string)
    
    # print(dict_string)
 
    # """read it back in"""
    # with open('data.json', 'r') as f:
    #     data_dict = json.load(f)

    # for k,v in data_dict.items():
    #     print(data_dict[k]['carbon emissions'])
    #     print(data_dict[k]['CPU %Utilization']['#Cores'])
    #     print(k,v)

    #     input("Press Enter to continue...")


    # Write the stringified flat dictionary to a file
    with open('data.txt', 'w') as f:
        f.write(dict_string)

    # Read the stringified flat dictionary from the file
    with open('data.txt', 'r') as f:
        read_back_string = f.read()

    print("Stringified flat dictionary read back from the file:")
    print(read_back_string)
